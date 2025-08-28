from pathlib import Path
from typing import List

from openpyxl import load_workbook


class ExcelReader:
    """用于读取Excel测试数据的工具类"""

    def __init__(self, file_name: str = "test_data.xlsx"):
        """
        初始化Excel读取器

        Args:
            file_name: Excel文件名，默认是test_data.xlsx
        """
        # 假设Excel文件位于项目根目录下的tests/data文件夹中
        self.base_path = Path(__file__).parent.parent.parent.parent / "tests" / "data"
        self.file_path = self.base_path / file_name

        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel文件不存在: {self.file_path}")

    def get_sheet_data_as_dict(self, sheet_name: str, skip_rows: int = 0) -> List[dict]:
        """
        获取指定工作表的数据，返回字典列表（第一行作为键）

        Args:
            sheet_name: 工作表名称
            skip_rows: 要跳过的行数，默认为0（不跳过任何行，从第一行读取标题）

        Returns:
            包含工作表数据的字典列表
        """
        workbook = load_workbook(self.file_path, read_only=True, data_only=True)

        try:
            sheet = workbook[sheet_name]

            # 获取标题行（跳过指定行数后的第一行）
            headers = []
            for cell in next(sheet.iter_rows(min_row=skip_rows + 1, max_row=skip_rows + 1, values_only=True)):
                headers.append(cell if cell is not None else f"Column_{len(headers) + 1}")

            # 读取数据行
            data = []
            for row in sheet.iter_rows(min_row=skip_rows + 2, values_only=True):
                # 跳过空行
                if any(cell is not None for cell in row):
                    row_data = dict(zip(headers, row))
                    data.append(row_data)

            workbook.close()
            return data

        except KeyError:
            available_sheets = workbook.sheetnames
            workbook.close()
            raise ValueError(
                f"工作表 '{sheet_name}' 不存在。可用的工作表: {available_sheets}"
            )
